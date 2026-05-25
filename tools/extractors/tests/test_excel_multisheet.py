"""Tests for the multi-sheet MITS Excel dictionary extractor."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from cora_extractors.config import ExcelMultiSheetDictionaryConfig
from cora_extractors.excel_multisheet import ExcelMultiSheetDictionaryExtractor


@pytest.fixture
def root_column_workbook(tmp_path: Path) -> Path:
    """A Lead-Management-style workbook: header at row 3 with `Common Root Element`."""
    wb = Workbook()
    wb.remove(wb.active)

    # Top sheet (will be ignored — not in type_sheets)
    top = wb.create_sheet("Module Top")
    top.append(["Category", "SubCategory", "Node", "Description"])
    top.append(["Module", "", "Foo", "module-level field"])

    # Type sheet with 3 hierarchy columns
    person = wb.create_sheet("Person Type")
    person.append(["Person Type", "", "", "", "", "", "", "", "", ""])
    person.append([""] * 10)
    person.append([""] * 10)
    person.append(
        [
            "Common Root Element",
            "Child Element/Attribute",
            "Child Element/Attribute",
            "Description",
            "Data Type",
            "Req",
            "Max occurs",
            "Restrictions",
            "Values",
        ]
    )
    # Row 4: type-level header (depth 0, root only) — skipped
    person.append(
        ["Person", "", "", "Person record", "Complex", "Y", "Unbounded", "", ""]
    )
    # Row 5: direct field on Person (depth 1)
    person.append(
        ["", "FirstName", "", "Person's first name", "xs:string", "Y", "1", "", ""]
    )
    # Row 6: another direct field with repeating cardinality
    person.append(
        ["", "Address", "", "Mailing address", "Complex", "N", "Unbounded", "", ""]
    )
    # Row 7: enumeration with comma-separated values
    person.append(
        [
            "",
            "Status",
            "",
            "Status of the record",
            "xs:string",
            "N",
            "1",
            "Enumerated values",
            "active, inactive, pending",
        ]
    )

    # Second type sheet with only 2 hierarchy columns
    rangetype = wb.create_sheet("NumericRangeType")
    rangetype.append(["NumericRangeType", "", "", "", "", "", "", "", ""])
    rangetype.append([""] * 9)
    rangetype.append([""] * 9)
    rangetype.append(
        [
            "Common Root Element",
            "Child Element/Attribute",
            "Description",
            "Data Type",
            "Req",
            "Max occurs",
            "Restrictions",
            "Values",
        ]
    )
    rangetype.append(
        ["NumericRangeType", "", "Range type header", "Complex", "Y", "1", "", ""]
    )
    rangetype.append(["", "Min", "Lower bound", "xs:integer", "N", "1", "", ""])
    rangetype.append(["", "Max", "Upper bound", "xs:integer", "N", "1", "", ""])

    path = tmp_path / "root-column.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def flat_workbook(tmp_path: Path) -> Path:
    """An Accounts-Payable-style workbook: header at row 0, no root column."""
    wb = Workbook()
    wb.remove(wb.active)

    addr = wb.create_sheet("Address Type")
    addr.append(
        [
            "Child Element/Attribute",
            "Child Element/Attribute",
            "Description",
            "Data Type",
            "Req",
            "Max occurs",
            "Restrictions",
            "Values",
        ]
    )
    # Row 1: type-name header inside data area — skipped because leaf == target_type
    addr.append(
        ["AddressType", "", "", "", "", "", "Enumerated List", "shipping, billing"]
    )
    # Row 2+: direct fields (depth 0 is valid in flat layout)
    addr.append(
        ["AddressLine1", "", "street address", "StringMax100", "No", "1", "", ""]
    )
    addr.append(["City", "", "city name", "StringMax60", "Yes", "1", "", ""])

    path = tmp_path / "flat.xlsx"
    wb.save(str(path))
    return path


def _config(type_sheets: dict[str, str]) -> ExcelMultiSheetDictionaryConfig:
    return ExcelMultiSheetDictionaryConfig(type_sheets=type_sheets)


# ---------------------------------------------------------------------------
# Root-column layout (Lead-Management-style)
# ---------------------------------------------------------------------------


def test_root_column_layout_extracts_fields_with_target_type_as_domain(
    root_column_workbook: Path,
) -> None:
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Person Type": "PersonType"})
    inv = extractor.extract(root_column_workbook, config, module="test")

    leaves = sorted(f.path.split("/")[-1] for f in inv.fields)
    assert leaves == ["Address", "FirstName", "Status"]
    for f in inv.fields:
        assert f.domain == "PersonType"
        assert f.path.startswith("PersonType/")


def test_root_column_layout_skips_depth_zero_rows(root_column_workbook: Path) -> None:
    """Row 4 (`Person` at col 0 only) is the type-context row and must be skipped."""
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Person Type": "PersonType"})
    inv = extractor.extract(root_column_workbook, config, module="test")
    assert not any(f.path.endswith("/Person") for f in inv.fields)


def test_root_column_layout_handles_variable_hierarchy_depth(
    root_column_workbook: Path,
) -> None:
    """NumericRangeType has 2 hierarchy columns; Person Type has 3 — both work."""
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config(
        {"Person Type": "PersonType", "NumericRangeType": "NumericRangeType"}
    )
    inv = extractor.extract(root_column_workbook, config, module="test")

    person_leaves = {f.path.split("/")[-1] for f in inv.fields if f.domain == "PersonType"}
    range_leaves = {
        f.path.split("/")[-1] for f in inv.fields if f.domain == "NumericRangeType"
    }
    assert person_leaves == {"FirstName", "Address", "Status"}
    assert range_leaves == {"Min", "Max"}


def test_root_column_layout_parses_enumeration_and_cardinality(
    root_column_workbook: Path,
) -> None:
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Person Type": "PersonType"})
    inv = extractor.extract(root_column_workbook, config, module="test")
    status = next(f for f in inv.fields if f.path.endswith("/Status"))
    assert status.enumeration == ["active", "inactive", "pending"]
    address = next(f for f in inv.fields if f.path.endswith("/Address"))
    assert address.cardinality == "repeating"  # Max occurs == Unbounded
    first_name = next(f for f in inv.fields if f.path.endswith("/FirstName"))
    assert first_name.cardinality == "required"  # Req == Y


# ---------------------------------------------------------------------------
# Flat layout (Accounts-Payable-style)
# ---------------------------------------------------------------------------


def test_flat_layout_emits_depth_zero_fields(flat_workbook: Path) -> None:
    """In flat layout (no root column), depth-0 rows ARE fields."""
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Address Type": "AddressType"})
    inv = extractor.extract(flat_workbook, config, module="test")

    leaves = sorted(f.path.split("/")[-1] for f in inv.fields)
    assert leaves == ["AddressLine1", "City"]


def test_flat_layout_skips_type_name_row(flat_workbook: Path) -> None:
    """The row whose leaf cell equals the target type name is type-level, skipped."""
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Address Type": "AddressType"})
    inv = extractor.extract(flat_workbook, config, module="test")
    assert not any(f.path.endswith("/AddressType") for f in inv.fields)


# ---------------------------------------------------------------------------
# Source-label discipline (required for downstream enrich)
# ---------------------------------------------------------------------------


def test_extractor_sets_excel_source_label(root_column_workbook: Path) -> None:
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Person Type": "PersonType"})
    inv = extractor.extract(root_column_workbook, config, module="test")
    assert inv.source_label == "excel"


# ---------------------------------------------------------------------------
# Error surfaces
# ---------------------------------------------------------------------------


def test_missing_sheet_raises(root_column_workbook: Path) -> None:
    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"Nonexistent Sheet": "Nope"})
    with pytest.raises(ValueError, match="not found"):
        extractor.extract(root_column_workbook, config, module="test")


def test_missing_header_marker_raises(tmp_path: Path) -> None:
    """If a configured sheet has no recognizable header, the extractor errors."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("BadSheet")
    ws.append(["nothing", "structured", "here"])
    ws.append(["random", "data", "row"])
    path = tmp_path / "bad.xlsx"
    wb.save(str(path))

    extractor = ExcelMultiSheetDictionaryExtractor()
    config = _config({"BadSheet": "Whatever"})
    with pytest.raises(ValueError, match="header row not found"):
        extractor.extract(path, config, module="test")


def test_rejects_wrong_config_type(root_column_workbook: Path) -> None:
    extractor = ExcelMultiSheetDictionaryExtractor()
    with pytest.raises(TypeError, match="ExcelMultiSheetDictionaryConfig"):
        extractor.extract(root_column_workbook, config=None, module="test")

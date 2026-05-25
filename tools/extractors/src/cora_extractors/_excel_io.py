"""Shared low-level helpers for Excel-based extractor adapters.

Module-private (underscored) — used by ``excel_dictionary`` (single-sheet
primary mode) and ``excel_multisheet`` (per-type sheet enrichment mode).
Both adapters share the same I/O and cell-coercion shape; the higher-level
"what does a row mean?" logic differs by mode and lives in each adapter.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, cast

from cora_extractors.inventory import Cardinality

VALID_CARDINALITY: frozenset[str] = frozenset({"required", "optional", "repeating"})


def column_to_index(letter: str) -> int:
    """Convert an Excel column letter (A, B, ..., AA, AB) to a 0-indexed number."""
    if not letter or not letter.isalpha():
        raise ValueError(f"invalid Excel column letter {letter!r}")
    n = 0
    for c in letter.upper():
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def read_rows(source: Path, sheet: str) -> list[list[Any]]:
    """Read every row of a sheet as a list of cell values."""
    suffix = source.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(filename=str(source), data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"sheet {sheet!r} not found in {source} (have: {wb.sheetnames})")
        ws = wb[sheet]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    if suffix == ".xls":
        import xlrd  # type: ignore[import-untyped]

        wb = xlrd.open_workbook(str(source))
        try:
            ws = wb.sheet_by_name(sheet)
        except xlrd.XLRDError as exc:
            raise ValueError(
                f"sheet {sheet!r} not found in {source} (have: {wb.sheet_names()})"
            ) from exc
        return [ws.row_values(i) for i in range(ws.nrows)]
    raise ValueError(f"unsupported Excel file extension: {suffix}")


def cell(row: list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def coerce_cardinality(raw: Any) -> Cardinality:
    """Coerce a free-form Excel value to a canonical Cardinality.

    Recognises:
    - Empty / None → 'optional'
    - Direct match for 'required' / 'optional' / 'repeating' → that value
    - 'Y'/'y' (in a Req column) → 'required'
    - 'N'/'n' (in a Req column) → 'optional'
    """
    if raw is None:
        return "optional"
    key = str(raw).strip()
    if not key:
        return "optional"
    lower = key.lower()
    if lower in VALID_CARDINALITY:
        return cast(Cardinality, lower)
    if lower in {"y", "yes"}:
        return "required"
    if lower in {"n", "no"}:
        return "optional"
    return "optional"


def parse_enumeration(raw: Any) -> list[str] | None:
    """Split an enumeration cell into a list.

    Excel data dictionaries vary on how enumerations are encoded. We split on
    common delimiters and let the config writer pre-clean if the source uses
    something exotic.
    """
    if raw is None or not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text:
        return None
    for delim in ("|", ";", ","):
        if delim in text:
            return [v.strip() for v in text.split(delim) if v.strip()]
    return [text]

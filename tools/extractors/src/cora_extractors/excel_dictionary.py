"""Excel data dictionary extractor adapter (primary mode only).

Each sheet in a data dictionary workbook becomes one inventory module.
Emits ``fields[]`` only — Excel is flat; ``types[]`` stays empty. To
enrich an XSD-derived inventory with definitions sourced from Excel,
extract both separately and use ``Inventory.merge(other, match_by='name')``.
"""

from __future__ import annotations

from collections.abc import Iterable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

from cora_extractors import __version__
from cora_extractors.config import ExcelDictionaryConfig, ExtractorConfig
from cora_extractors.inventory import Cardinality, FieldEntry, Inventory
from cora_extractors.path import build as build_path

EXTRACTOR_ID = f"cora_extractors.excel_dictionary@{__version__}"

_VALID_CARDINALITY = {"required", "optional", "repeating"}


def _column_to_index(letter: str) -> int:
    """Convert a column letter (A, B, ..., AA, AB) to a 0-indexed column number."""
    if not letter or not letter.isalpha():
        raise ValueError(f"invalid Excel column letter {letter!r}")
    n = 0
    for c in letter.upper():
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def _read_rows(source: Path, sheet: str) -> list[list[Any]]:
    """Read every row of a sheet as a list of cell values."""
    suffix = source.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(filename=str(source), data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"sheet {sheet!r} not found in {source} (have: {wb.sheetnames})")
        ws = wb[sheet]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    if suffix == ".xls":
        import xlrd  # type: ignore[import-untyped]

        wb = xlrd.open_workbook(str(source))
        try:
            ws = wb.sheet_by_name(sheet)
        except xlrd.XLRDError as exc:
            raise ValueError(
                f"sheet {sheet!r} not found in {source} (have: {wb.sheet_names()})"
            ) from exc
        return [ws.row_values(i) for i in range(ws.nrows)]
    raise ValueError(f"unsupported Excel file extension: {suffix}")


class ExcelDictionaryExtractor:
    """Extractor adapter for Excel data dictionaries."""

    name = "excel"

    def extract(
        self,
        source: Path,
        config: ExtractorConfig | None = None,
        *,
        module: str | None = None,
    ) -> Inventory:
        if not isinstance(config, ExcelDictionaryConfig):
            raise TypeError(
                "ExcelDictionaryExtractor.extract requires an ExcelDictionaryConfig"
            )

        rows = _read_rows(source, config.sheet)
        body = rows[config.skip_rows :]
        fields = list(_iter_fields(body, config, source.name))

        return Inventory(
            standard=module or source.stem,
            module=module or source.stem,
            version="unknown",
            source_artifact=str(source),
            extractor=EXTRACTOR_ID,
            extracted_at=datetime.now(tz=UTC),
            namespace_hint=config.namespace_hint,
            types=[],
            fields=fields,
        )


def _iter_fields(
    rows: list[list[Any]], config: ExcelDictionaryConfig, source_name: str
) -> Iterable[FieldEntry]:
    name_idx = _column_to_index(config.field_name_column)
    type_idx = _column_to_index(config.type_column) if config.type_column else None
    def_idx = _column_to_index(config.definition_column) if config.definition_column else None
    card_idx = (
        _column_to_index(config.cardinality_column) if config.cardinality_column else None
    )
    enum_idx = (
        _column_to_index(config.enumeration_column) if config.enumeration_column else None
    )

    for row_offset, row in enumerate(rows):
        name = _cell(row, name_idx)
        if not isinstance(name, str) or not name.strip():
            continue
        clean_name = name.strip()
        path = build_path([*config.path_prefix, clean_name])
        raw_type = _cell(row, type_idx) if type_idx is not None else None
        raw_def = _cell(row, def_idx) if def_idx is not None else None
        raw_card = _cell(row, card_idx) if card_idx is not None else None
        raw_enum = _cell(row, enum_idx) if enum_idx is not None else None

        # Excel rows are 1-indexed for humans; +1 for header skip recorded separately.
        sheet_row = config.skip_rows + row_offset + 1
        source_location = f"{source_name}!{config.sheet}!{sheet_row}"

        yield FieldEntry(
            path=path,
            range=str(raw_type).strip() if isinstance(raw_type, str) and raw_type else None,
            cardinality=_coerce_cardinality(raw_card, config),
            definition=str(raw_def).strip() if isinstance(raw_def, str) else "",
            source_location=source_location,
            enumeration=_parse_enumeration(raw_enum),
        )


def _cell(row: list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _coerce_cardinality(raw: Any, config: ExcelDictionaryConfig) -> Cardinality:
    if raw is None:
        return "optional"
    key = str(raw).strip()
    if not key:
        return "optional"
    if key in _VALID_CARDINALITY:
        return cast(Cardinality, key)
    return "optional"


def _parse_enumeration(raw: Any) -> list[str] | None:
    """Split an enumeration cell into a list.

    Excel data dictionaries vary wildly on how enumerations are encoded. The
    safest default is to split a string on common delimiters and let the
    config writer pre-clean if the source uses something exotic.
    """
    if raw is None or not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text:
        return None
    for delim in ("|", ";", ","):
        if delim in text:
            return [v.strip() for v in text.split(delim) if v.strip()]
    return [text]
